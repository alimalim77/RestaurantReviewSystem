# import streamlit as st
# import pandas as pd
# import plotly.express as px
# import openpyxl as op
# import webbrowser

# # Set page configuration
# st.set_page_config(page_title="Customer Reviews")

# with open("style.css") as f:
#     st.markdown("<style>{}</style>".format(f.read()), unsafe_allow_html=True)

# st.header("Analyzed Customer Reviews")
# st.subheader("Ratings from the customer")

# excel_file = 'Reviews_Ans.xlsx'
# excel_file2 = 'opinion.xlsx'
# sheet = 'Sheet1'

# # Takes the following total columns and presents them
# st.header("Customer Information")
# df = pd.read_excel(excel_file,
#                    sheet_name=sheet,
#                    usecols='A:C',
#                    header=0,)

# # Take the opinion excel file and generates a pie chart
# st.subheader("Information on the reviews informations")
# opinion = pd.read_excel(excel_file2,
#                         sheet_name=sheet,
#                         usecols='A:C',
#                         header=0,)
# workbook_op = op.load_workbook('opinion.xlsx')
# sheet_op = workbook_op.active
# good = sheet_op.cell(row=2, column=1).value
# bad = sheet_op.cell(row=2, column=2).value


# pie = px.pie(title="Reviews", values=[good, bad], names=["Good", "Bad"])
# workbook_op.save('opinion.xlsx')

# link = "https://t.me/myRestaurant123_bot"
# button_label = "Go to Telegram"

# if st.button(button_label):
#     webbrowser.open_new_tab(link)

# st.image("telegram.png", "Send a message to users", 100)

# st.dataframe(df)
# st.dataframe(opinion)
# st.plotly_chart(pie)
import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl as op
import webbrowser
from pathlib import Path
import pickle
import streamlit_authenticator as stauth

# Set page configuration
st.set_page_config(page_title="Customer Reviews")

names = ["Alim", "Sahil", "Devashish"]
usernames = ["ma", "msk", "dk"]
password = ["AM", "KSM", "KD"]

# load hashed passwords
file_path = Path(__file__).parent / "hashed_pw.pkl"
with file_path.open("rb") as file:
    hashed_passwords = pickle.load(file)

credentials = {"usernames": {}}

for un, name, pw in zip(usernames, names, hashed_passwords):
    user_dict = {"name": name, "password": pw}
    credentials["usernames"].update({un: user_dict})

authenticator = stauth.Authenticate(
    credentials, "Customer Reviews", "abcdef", 30)
name, authentication, username = authenticator.login("Login", "main")


if authentication == False:
    st.error("Username/password is incorrect")

if authentication == None:
    st.warning("Please enter your username and password")

if authentication:
    with open("style.css") as f:
        st.markdown("<style>{}</style>".format(f.read()),
                    unsafe_allow_html=True)

    # Wrap all content in a container div
    st.markdown(
        """
        <div style='display: flex; flex-direction: column; align-items: center;'>
        """, unsafe_allow_html=True)

    st.header("Analyzed Customer Reviews")
    st.subheader("Ratings from the customer")

    excel_file = 'Reviews_Ans.xlsx'
    excel_file2 = 'opinion.xlsx'
    sheet = 'Sheet1'

    # Takes the following total columns and presents them
    st.header("Customer Information")
    df = pd.read_excel(excel_file,
                       sheet_name=sheet,
                       usecols='A:C',
                       header=0,)
    authenticator.logout("Logout", "sidebar")
    st.sidebar.title(f"Welcome {name}")
    # Take the opinion excel file and generates a pie chart
    st.subheader("Information on the reviews informations")
    opinion = pd.read_excel(excel_file2,
                            sheet_name=sheet,
                            usecols='A:C',
                            header=0,)
    workbook_op = op.load_workbook('opinion.xlsx')
    sheet_op = workbook_op.active
    good = sheet_op.cell(row=2, column=1).value
    bad = sheet_op.cell(row=2, column=2).value

    pie = px.pie(title="Reviews", values=[good, bad], names=["Good", "Bad"])
    workbook_op.save('opinion.xlsx')

    link = "https://t.me/myRestaurant123_bot"
    button_label = "Go to Telegram"

    if st.button(button_label):
        webbrowser.open_new_tab(link)

    st.image("telegram.png", "Send a message to users", 100)

    st.dataframe(df)
    st.dataframe(opinion)
    st.plotly_chart(pie)

    # Close the container div
    st.markdown("</div>", unsafe_allow_html=True)
