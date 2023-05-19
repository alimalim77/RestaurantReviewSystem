import telebot as tl
import analysis
import openpyxl as op
import datetime

current_date = datetime.datetime.now()
food = ("biryani", "pizza", "waffle", "pasta", "fries")

BOT_TOKEN = '6127391390:AAF2yZ1L828iUyROP6wyhZaTz2Odfq77ggU'

bot = tl.TeleBot(BOT_TOKEN)
fdbk = ''
score_log = []


@bot.message_handler(commands=['start', 'hello'])
def send_welcome(message):
    bot.reply_to(
        message, "Hello!! we hope yoy had a wonderful exp. dining in here, we would love to hear it from you.")
    print(dir(message), type(message))
    print(message)
    print(message.chat.id, message.chat.first_name, message.chat.last_name)


@bot.message_handler(func=lambda msg: True)
def echo_all(message):
    global fdbk
    global score_log
    # name = message.name
    d_time = message.date
    fid = message.chat.id
    fdbk = message.text
    score = analysis.sia(fdbk)

    flag = False
    fd_items = ""
    for i in fdbk.split():
        if i.lower() in food:
            fd_items = i.lower()
            flag = True
    # set up for creating excel file
    workbook = op.load_workbook('Reviews_Ans.xlsx')
    workbook_op = op.load_workbook('opinion.xlsx')

    sheet = workbook.active
    sheet_op = workbook_op.active
    score_log.append(score)
    if flag:
        row = [fid, fdbk, score, fd_items,
               datetime.datetime.fromtimestamp(d_time)]
    else:
        row = [fid, fdbk, score, "NULL",
               datetime.datetime.fromtimestamp(d_time)]
    row = [fid, fdbk, score]
    sheet.append(row)
    workbook.save('Reviews_Ans.xlsx')

    # Get the current values of the cells in the second row
    good = sheet_op.cell(row=2, column=1).value
    bad = sheet_op.cell(row=2, column=2).value

    # response to user based on score
    if score < -0.2:
        bot.reply_to(message, "Apologies for the subpar experience.")
        bad += 1
    elif score > 0.2:
        good += 1
        bot.reply_to(message, "Pleased to serve you well!")
        bot.reply_to(message, "Thanks for choosing us!! Have a Great Day.")
    else:
        bot.reply_to(message, "Thanks for choosing us!!")

    sheet_op.cell(row=2, column=1, value=good)
    sheet_op.cell(row=2, column=2, value=bad)
    workbook_op.save('opinion.xlsx')

    print(message.text)
    print(message)


bot.infinity_polling()
